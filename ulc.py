# expected input: excel invoice, on sheet no. 1. All dates onf column2 should be 
# formatted as dd/mm/yyyy. 

import os
import csv
import openpyxl as exl
import easygui
from datetime import datetime

def RepresentsInt(s):
    if s is None:
        return False
    try: 
        int(s)
        return True
    except ValueError:
        return False

def main():
    fields=['SupplierCode','AccountNumber','InvoiceDate','InvoiceNumber','OriginalAmountDue'\
    ,'TrackingNumber','TansportationCharge','NetChargeAmount','ServiceType','ShipmentDate'\
    ,'ActualWeightAmount','RatedWeightAmount','RatedWeightUnits','NumberofPieces','RecipientCompany'\
    ,'RecipientZipCode','RecipientCountry','ShipperZipCode','ShipperCountry','CustomerReference'\
    ,'CustomerReference#2','ZoneCode','OrderCharge1','OrderChargeAmount1','OrderCharge2','OrderChargeAmount2']

    ulc_charges_mapping={5:'CCL',6:'MIS',7:'IET',8:'PHI',9:'ACC3',10:'EPD'}
    invoice_num=input('Invoice number:')
    # invoice_num='4'
    # account_num=input('Account number:')
    account_num=None
    invoice_date= input('Invoice date(yyyymmdd):')
    # invoice_date='20190913'
    outdir = os.path.dirname("outfiles/")
    os.makedirs(outdir,exist_ok=True)

    dest_rows=list()
    invoice_file_name=easygui.fileopenbox(msg="select excel invoice to upload",)
    if invoice_file_name is None:
        exit()
    try:
        wb = exl.load_workbook(filename = invoice_file_name,data_only=True)
        active_sheet=wb[wb.sheetnames[0]]
    except:
        raise Exception("Problem reading invoice file")

    # rownum=1
    total_amount=0
    entities=['noentity']*active_sheet.max_row
    for row in active_sheet.iter_rows(2,active_sheet.max_row,1,20):
        rownum+=1
        for column in range(5,11):
            charge_amount=row[column].value
            operation_date=row[1].value   
            if RepresentsInt(charge_amount) and operation_date is not None:
                dest_row={}.fromkeys(fields,None)
                dest_row['SupplierCode']='UL'
                dest_row['AccountNumber']=account_num
                total_amount+=charge_amount
                dest_row['InvoiceDate']=invoice_date
                dest_row['InvoiceNumber']=invoice_num
                dest_row['TrackingNumber']=0
                dest_row['TansportationCharge']=0
                dest_row['NetChargeAmount']=round(charge_amount,2)
                dest_row['OrderChargeAmount1']=round(charge_amount,2)
                service_type=ulc_charges_mapping.get(column)
                dest_row['ServiceType']=service_type
                dest_row['OrderCharge1']=service_type
                dest_row['ShipmentDate']=operation_date.strftime('%Y%m%d')
                dest_row['NumberofPieces']=1
                dest_row['RecipientCompany']='ULC'
                if entities[rownum]!='noentity' :
                    entity_num=entities[rownum] 
                else:
                    entity_num=str(row[3].value).strip()
                while True:
                    if entity_num=='none':
                        entities[rownum]=entity_num
                        break                
                    if RepresentsInt(entity_num):
                        if len(str(entity_num))==6:
                            if str(entity_num)[0]=='2':
                                dest_row['CustomerReference#2']=entity_num
                                if entities[rownum]=='noentity':
                                    entities[rownum]=entity_num
                                break
                            elif str(entity_num)[0]=='4':
                                dest_row['CustomerReference']=entity_num  
                                if entities[rownum]=='noentity':
                                    entities[rownum]=entity_num
                                break                        
                    tracking_number=row[2].value
                    entity_num= input("Please enter entity number[Tracking id:"+str(tracking_number)+\
                    ". Operation date:"+operation_date.strftime('%Y%m%d')+' Customer:'+row[4].value+\
                        "]. Type 'none' for no entity:  ")
                dest_rows.append(dest_row)

    for v in dest_rows:
        v['OriginalAmountDue']=round(total_amount,2)
    outfile_name='outfiles\\'+str(invoice_num) +'-inbound-'+ datetime.now().strftime('%Y%m%d%H%M%S')+'.txt'

    with open(outfile_name, 'w',newline='')as outcsv:
        w= csv.DictWriter(outcsv,fields,delimiter='\t')
        w.writeheader()
        w.writerows(dest_rows)
    #remove last line
    with open(outfile_name, 'r') as f:
        data = f.read().rstrip('\n')
    with open(outfile_name, 'w') as f_output:    
        f_output.write(data)