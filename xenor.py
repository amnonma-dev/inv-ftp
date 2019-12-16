import os
import csv
import openpyxl as exl
import easygui
from datetime import datetime
from dateutil.parser import parse


def RepresentsInt(s):
    if s is None:
        return False
    try:
        int(s)
        return True
    except ValueError:
        return False


def main():
    fields = ['SupplierCode', 'AccountNumber', 'InvoiceDate', 'InvoiceNumber', 'OriginalAmountDue', 'TrackingNumber', 'TansportationCharge', 'NetChargeAmount', 'ServiceType', 'ShipmentDate', 'ActualWeightAmount', 'RatedWeightAmount', 'RatedWeightUnits',
              'NumberofPieces', 'RecipientCompany', 'RecipientZipCode', 'RecipientCountry', 'ShipperZipCode', 'ShipperCountry', 'CustomerReference', 'CustomerReference#2', 'ZoneCode', 'OrderCharge1', 'OrderChargeAmount1', 'OrderCharge2', 'OrderChargeAmount2']
    upload_type = input(
        "press 'i' for inbound, 'o' for outbound, 'q' for quit:")
    while upload_type != 'i' and upload_type != 'o' and upload_type != 'q':
        upload_type = input(
            "press 'i' for inbound, 'o' for outbound, 'q' for quit")
    if upload_type == 'q':
        exit()

    invoice_num = input('Invoice number:')
    # invoice_num='5131960142'
    # account_num=input('Account number:')
    account_num = '12150149'
    invoice_date = input('Invoice date(yyyymmdd):')
    # invoice_date='20190913'
    outdir = os.path.dirname("outfiles/")
    os.makedirs(outdir, exist_ok=True)

    invoice_file_name = easygui.fileopenbox(
        msg="select  excel invoice to upload",)
    # invoice_file_name='xenor.xlsx'
    if invoice_file_name is None:
        exit()
    try:
        wb = exl.load_workbook(filename=invoice_file_name, data_only=True)
        active_sheet = wb[wb.sheetnames[0]]
    except:
        raise Exception("Problem reading invoice file")

    total_amount = 0
    dest_rows = list()
    for row in active_sheet.iter_rows(1, active_sheet.max_row, 1, 20):
        dest_row = {}.fromkeys(fields, None)
        if upload_type == 'i':
            operation_date = parse(row[1].value)
            dest_row['ServiceType'] = "IHL"
            dest_row['OrderCharge1'] = "IHL"
            tracking_number = row[4].value
            dest_row['NumberofPieces'] = row[3].value
            if len(str(tracking_number)) == 1:
                tracking_number = row[5].value
            entity_num = row[10].value
            customer=row[2].value
            type_of_activity = row[6].value
            if type_of_activity == 'Replenishment':
                charge_amount = row[3].value * 0.3
                total_amount = total_amount+charge_amount
                charge_amount = round(charge_amount, 2)

            elif type_of_activity == 'RMA InBound' or type_of_activity == 'Return To Stock':
                charge_amount = row[3].value * 0.9
                total_amount = total_amount+charge_amount
                charge_amount = round(charge_amount, 2)
            else:
                continue
        elif upload_type == 'o':
            # print (row[9].value)
            operation_date = parse(row[9].value)
            dest_row['ServiceType'] = "OHL"
            dest_row['OrderCharge1'] = "OHL"
            tracking_number = row[2].value
            dest_row['NumberofPieces'] = row[6].value
            customer=row[0].value
            entity_num = row[3].value
            charge_amount = row[6].value * 0.35 + 1
            charge_amount = round(charge_amount, 2)
            total_amount = total_amount+charge_amount
        dest_row['NetChargeAmount'] = charge_amount
        dest_row['OrderChargeAmount1'] = charge_amount

        dest_row['ShipmentDate'] = operation_date.strftime('%Y%m%d')
        dest_row['SupplierCode'] = 'XNR'
        dest_row['AccountNumber'] = account_num
        dest_row['InvoiceDate'] = invoice_date
        dest_row['InvoiceNumber'] = invoice_num
        dest_row['TrackingNumber'] = tracking_number
        # print('row num:'+str(rownum)+'\ttracking:'+dest_row['TrackingNumber'])
        dest_row['TansportationCharge'] = 0
        dest_row['RecipientCompany'] = 'Xenor Logistics'
        while True:
            if RepresentsInt(entity_num):
                entity_num = int(entity_num)
                if upload_type=='i':
                    if entity_num >= 200000 and entity_num <= 350000:
                        break
                elif upload_type=='o':
                    if entity_num >= 400000 and entity_num <= 550000:
                        break

            entity_num = input("Please enter entity number[Tracking id:"+str(tracking_number) +
                               ". Operation date:"+operation_date.strftime('%Y%m%d')+' Customer:'+str(customer) +
                               "]. Type 'none' for no entity:  ")
            if entity_num == 'none':
                break
        if upload_type=='i':
            dest_row['CustomerReference#2'] = entity_num
        elif upload_type=='o':        
            dest_row['CustomerReference'] = entity_num
        dest_rows.append(dest_row)
    # print(round(total_amount,2))
    for v in dest_rows:
        v['OriginalAmountDue'] = round(total_amount, 2)
    outfile_name = 'outfiles\\' + \
        str(invoice_num) + '-'+upload_type+'-' + \
        datetime.now().strftime('%Y%m%d%H%M%S')+'.txt'

    with open(outfile_name, 'w', newline='',encoding="utf-8")as outcsv:
        w = csv.DictWriter(outcsv, fields, delimiter='\t')
        w.writeheader()
        w.writerows(dest_rows)
    # remove last line
    with open(outfile_name, 'r') as f:
        data = f.read().rstrip('\n')
    with open(outfile_name, 'w') as f_output:
        f_output.write(data)

