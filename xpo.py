import os
import csv
import openpyxl as exl
import easygui
from datetime import datetime



def RepresentsInt(s):
    if s is None:
        return False
    try: 
        int(s)
        return True
    except ValueError:
        return False

def get_service_type(service_type,charges):
    for k,v in charges.items():
        if k.lower()==service_type.lower():
            return v

def main():
    fields=['SupplierCode','AccountNumber','InvoiceDate','InvoiceNumber','OriginalAmountDue'\
    ,'TrackingNumber','TansportationCharge','NetChargeAmount','ServiceType','ShipmentDate'\
    ,'ActualWeightAmount','RatedWeightAmount','RatedWeightUnits','NumberofPieces','RecipientCompany'\
    ,'RecipientZipCode','RecipientCountry','ShipperZipCode','ShipperCountry','CustomerReference'\
    ,'CustomerReference#2','ZoneCode','OrderCharge1','OrderChargeAmount1','OrderCharge2','OrderChargeAmount2']

    invoice_num=input('Invoice number:')
    # invoice_num='5131960142'
    # account_num=input('Account number:')
    account_num='702169343'
    invoice_date= input('Invoice date(yyyymmdd):')
    # invoice_date='20190913'
    outdir = os.path.dirname("outfiles/")
    os.makedirs(outdir,exist_ok=True)

    charge_types={}
    with open ('charge-types.txt','r') as f:
        for line in f:
            charge_types[line.split('\t')[0].title()]= line.split('\t')[1]
    dest_rows=list()
    invoice_file_name=easygui.fileopenbox(msg="select excel invoice to upload",)
    if invoice_file_name is None:
        exit()
    try:
        wb = exl.load_workbook(filename = invoice_file_name,data_only=True)
        inbound_sheet=wb[wb.sheetnames[1]]
        outbound_sheet=wb[wb.sheetnames[2]]
    except:
        raise Exception("Problem reading invoice file")
    #******** INBOUND *******************

    rownum=1
    total_amount=0
    for row in inbound_sheet.iter_rows(2,inbound_sheet.max_row,1,20):
        dest_row={}.fromkeys(fields,None)
        rownum+=1
        charge_amount=row[10].value
        operation_date=row[6].value
        if charge_amount and operation_date and charge_amount>0:
            dest_row['SupplierCode']='MHK'
            dest_row['AccountNumber']=account_num
            total_amount+=charge_amount
            dest_row['InvoiceDate']=invoice_date
            dest_row['InvoiceNumber']=invoice_num
            tracking_number= row[14].value
            
            sign_index=-1
            try:
                sign_index=tracking_number.index('#')
            except:
                pass
            if tracking_number:
                dest_row['TrackingNumber']=tracking_number[sign_index+1:]
            # print('row num:'+str(rownum)+'\ttracking:'+dest_row['TrackingNumber'])
            dest_row['TansportationCharge']=0
            dest_row['NetChargeAmount']=round(charge_amount,2)
            dest_row['OrderChargeAmount1']=round(charge_amount,2)
            service_type=row[12].value
            if service_type:
                dest_row['ServiceType']=get_service_type(service_type,charge_types)
                dest_row['OrderCharge1']=get_service_type(service_type,charge_types)
            dest_row['ShipmentDate']=operation_date.strftime('%Y%m%d')
            dest_row['NumberofPieces']=row[7].value
            dest_row['RecipientCompany']='XPO HK'
            entity_num=row[5].value
            while True:
                try:
                    entity_num=int(entity_num)
                    if len(str(entity_num))==6:
                        break
                    else:
                        raise Exception('')
                except:
                    entity_num= input("Please enter entity number[Tracking id:"+str(tracking_number)+\
                    ". Operation date:"+operation_date.strftime('%Y%m%d')+' Customer:'+row[3].value+\
                        "]. Type 'none' for no entity:  ")
                    if entity_num=='none':
                        break
                
            if entity_num!='none':
                if entity_num>400000:
                    dest_row['CustomerReference']=entity_num
                else:
                    dest_row['CustomerReference#2']=entity_num
            dest_rows.append(dest_row)

    for v in dest_rows:
        v['OriginalAmountDue']=round(total_amount,2)
    outfile_name='outfiles\\'+str(invoice_num) +'-inbound-'+ datetime.now().strftime('%Y%m%d%H%M%S')+'.txt'

    with open(outfile_name, 'w',newline='')as outcsv:
        w= csv.DictWriter(outcsv,fields,delimiter='\t')
        w.writeheader()
        w.writerows(dest_rows)
    #remove last line
    with open(outfile_name, 'r') as f:
        data = f.read().rstrip('\n')
    with open(outfile_name, 'w') as f_output:    
        f_output.write(data)
        
    #******** OUTBOUND *******************

    rownum=1
    total_amount=0

    dest_rows=list()
    for row in outbound_sheet.iter_rows(2,outbound_sheet.max_row,1,20):
        dest_row={}.fromkeys(fields,None)
        rownum+=1
        charge_amount=row[10].value
        operation_date=row[6].value
        if RepresentsInt(charge_amount) and operation_date and charge_amount>0:
            dest_row['SupplierCode']='MHK'
            dest_row['AccountNumber']=account_num
            total_amount+=charge_amount
            dest_row['InvoiceDate']=invoice_date
            dest_row['InvoiceNumber']=invoice_num
            tracking_number= row[14].value
            
            sign_index=-1
            try:
                sign_index=tracking_number.index('#')
            except:
                pass
            if tracking_number:
                dest_row['TrackingNumber']=tracking_number[sign_index+1:]
            # print('row num:'+str(rownum)+'\ttracking:'+dest_row['TrackingNumber'])
            dest_row['TansportationCharge']=0
            dest_row['NetChargeAmount']=round(charge_amount,2)
            dest_row['OrderChargeAmount1']=round(charge_amount,2)
            service_type=row[12].value
            if service_type:
                dest_service_type=get_service_type(service_type,charge_types)
                if dest_service_type=="IHL":
                    dest_service_type="OHL"
                dest_row['ServiceType']=dest_service_type
                dest_row['OrderCharge1']=dest_service_type
            dest_row['ShipmentDate']=operation_date.strftime('%Y%m%d')
            dest_row['NumberofPieces']=row[7].value
            dest_row['RecipientCompany']='XPO HK'
            entity_num=row[4].value
            while True:
                try:
                    entity_num=int(entity_num)
                    if len(str(entity_num))==6 and entity_num>=400000:
                        break
                    else:
                        raise Exception('')
                except:
                    entity_num= input("Please enter OC number[Tracking id:"+str(tracking_number)+\
                    ". Operation date:"+operation_date.strftime('%Y%m%d')+' Customer:'+row[2].value+\
                        "]. Type 'none' for no entity:  ")
                    if entity_num=='none':
                        break
                
            if entity_num!='none':
                dest_row['CustomerReference']=entity_num
            dest_rows.append(dest_row)

    for v in dest_rows:
        v['OriginalAmountDue']=round(total_amount,2)
    outfile_name='outfiles\\'+str(invoice_num) +'-outbound-'+ datetime.now().strftime('%Y%m%d%H%M%S')+'.txt'

    with open(outfile_name, 'w',newline='')as outcsv:
        w= csv.DictWriter(outcsv,fields,delimiter='\t')
        w.writeheader()
        w.writerows(dest_rows)
    #remove last line
    with open(outfile_name, 'r') as f:
        data = f.read().rstrip('\n')
    with open(outfile_name, 'w') as f_output:    
        f_output.write(data)